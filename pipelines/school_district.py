"""国土数値情報 A27 小学校区 / A32 中学校区データのダウンロード。

国土交通省「国土数値情報ダウンロードサイト」から通学区域のポリゴンを
都道府県単位でダウンロードし、都道府県ごとの Parquet に変換して配置する。

利用条件による絞り込み
----------------------
通学区域データは原典資料を作成した市区町村ごとに公開条件が異なる。提供元が
配布する「公開に関する利用条件」（市区町村別の一覧）には、公開条件（公開 /
公開不可 / 作成不可）と利用条件（オープンデータ公開 / 条件公開 / 商用利用不可 /
商用利用・再配布不可）が小学校区・中学校区それぞれについて示されている。
配布 zip にはこれらが区別なく入っているため、利用条件が「オープンデータ公開」の
市区町村だけをここで取り込む。判定は利用条件一覧を毎回取得して行い、zip に含まれる
市区町村が一覧に載っていない場合は判定できないためエラーで停止する。

小学校区と中学校区は同じ市区町村でも利用条件が異なることがあるので、
一覧の列も別々に見る。

データソース: 小学校区データ（A27、2023年度）/ 中学校区データ（A32、2023年度）
https://nlftp.mlit.go.jp/ksj/gml/datalist/KsjTmplt-A27-2023.html
https://nlftp.mlit.go.jp/ksj/gml/datalist/KsjTmplt-A32-2023.html
"""

import logging
import os
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path
from urllib.error import HTTPError
from urllib.request import Request, urlopen

import duckdb
import openpyxl

logger = logging.getLogger("pipelines")

# 市区町村別の公開に関する利用条件（提供元が配布する Excel。小中学校区の共通ファイル）
TERMS_URL = (
    "https://nlftp.mlit.go.jp/ksj/gml/codelist/R5_Terms_of_use_municipality_data.xlsx"
)

# 一覧の列位置（0 始まり）と、その列に入っているはずの見出し
TERMS_COLUMNS = {
    "admin_code": (1, "行政CD"),
    "municipality_name": (2, "自治体"),
    "elementary_disclosure": (3, "小学校区"),
    "elementary_terms": (4, "小学校区"),
    "junior_high_disclosure": (7, "中学校区"),
    "junior_high_terms": (8, "中学校区"),
}

# 取り込む利用条件。これ以外（条件公開・商用利用不可・商用利用・再配布不可）は除外する
OPEN_TERMS = "オープンデータ公開"

# 一覧のパース結果が壊れていないことを確かめる下限
# （令和5年度版は 1,621 市区町村・うちオープンデータ公開 小 1,564 / 中 1,566）
MIN_TERMS_ROWS = 1500
MIN_OPEN_ROWS = 1400

PREFECTURES = [f"{i:02d}" for i in range(1, 48)]


@dataclass(frozen=True)
class SchoolDistrict:
    """通学区域データの種別ごとの取得条件。"""

    key: str  # 出力ディレクトリ名
    identifier: str  # 国土数値情報の識別子つきファイル名の接頭辞（例 A27-23）
    attribute_prefix: str  # GeoJSON の属性名の接頭辞（例 A27）
    terms_column: str  # 利用条件一覧のどの列を見るか

    @property
    def url_template(self) -> str:
        code = self.identifier.split("-")[0]
        return (
            f"https://nlftp.mlit.go.jp/ksj/gml/data/{code}/{self.identifier}/"
            f"{self.identifier}_{{pref}}_GML.zip"
        )


DISTRICTS = [
    SchoolDistrict("elementary", "A27-23", "A27", "elementary_terms"),
    SchoolDistrict("junior_high", "A32-23", "A32", "junior_high_terms"),
]

CREATE_TABLE_SQL = """
CREATE OR REPLACE TEMP TABLE school_district (
    admin_code VARCHAR,
    establisher VARCHAR,
    school_code VARCHAR,
    school_name VARCHAR,
    address VARCHAR,
    geom BLOB
)
"""


def _prefectures() -> list[str]:
    """処理対象の都道府県コード一覧。

    NLFTP_SCHOOL_DISTRICT_PREFECTURES（カンマ区切り、例: "11,13"）で絞り込める。
    未指定なら全国（01〜47）。
    """
    env = os.environ.get("NLFTP_SCHOOL_DISTRICT_PREFECTURES")
    if env:
        return [p.strip() for p in env.split(",") if p.strip()]
    return PREFECTURES


def _download(url: str, dest: Path) -> None:
    req = Request(url, headers={"User-Agent": "dataset-nlftp"})
    with urlopen(req) as resp, open(dest, "wb") as f:
        shutil.copyfileobj(resp, f, 1024 * 1024)


def _decode_member_name(name: str) -> str:
    """zip 内のファイル名を CP932 として復元する。

    zip にはファイル名が CP932 の日本語ファイルが含まれるため、
    Python zipfile が CP437 として解釈した文字列を元に戻す。
    復元できない場合はそのまま返す。
    """
    try:
        return name.encode("cp437").decode("cp932")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def _parse_terms(xlsx_path: Path) -> list[dict[str, str]]:
    """利用条件一覧を市区町村ごとの辞書のリストにする。

    列の並びが変わったまま読み進めると誤った市区町村を公開してしまうため、
    見出し行が想定どおりかを先に確かめる。
    """
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header = rows[0]
    for field, (index, expected) in TERMS_COLUMNS.items():
        actual = str(header[index] or "").replace("\n", "")
        if expected not in actual:
            raise SystemExit(
                f"terms sheet layout changed: column {index} for {field} "
                f"is {actual!r}, expected to contain {expected!r}"
            )

    terms = []
    for row in rows[1:]:
        if not row[TERMS_COLUMNS["admin_code"][0]]:
            continue
        terms.append(
            {
                field: str(row[index] or "")
                for field, (index, _) in TERMS_COLUMNS.items()
            }
        )
    return terms


def _load_terms(dest: Path) -> list[dict[str, str]]:
    """利用条件一覧を取得してパースする。"""
    xlsx_path = dest / "terms_of_use.xlsx"
    logger.info("  downloading terms of use...")
    _download(TERMS_URL, xlsx_path)
    try:
        terms = _parse_terms(xlsx_path)
    finally:
        xlsx_path.unlink(missing_ok=True)

    counts = {
        district.key: len(_open_codes(terms, district)) for district in DISTRICTS
    }
    if len(terms) < MIN_TERMS_ROWS or min(counts.values()) < MIN_OPEN_ROWS:
        raise SystemExit(
            f"terms of use parse looks broken: {len(terms)} municipalities, "
            f"open data {counts}"
        )

    logger.info(f"  terms of use: {len(terms)} municipalities, open data {counts}")
    return terms


def _open_codes(terms: list[dict[str, str]], district: SchoolDistrict) -> set[str]:
    """利用条件が「オープンデータ公開」の市区町村コード。"""
    return {t["admin_code"] for t in terms if t[district.terms_column] == OPEN_TERMS}


def _write_terms(dest: Path, terms: list[dict[str, str]]) -> None:
    """市区町村別の公開条件・利用条件を Parquet に書き出す。"""
    columns = list(TERMS_COLUMNS)
    con = duckdb.connect()
    try:
        con.execute(
            "CREATE TEMP TABLE terms ("
            + ", ".join(f"{name} VARCHAR" for name in columns)
            + ")"
        )
        con.executemany(
            f"INSERT INTO terms VALUES ({', '.join('?' * len(columns))})",
            [
                tuple(t[name] for name in columns)
                for t in sorted(terms, key=lambda t: t["admin_code"])
            ],
        )
        con.execute(
            f"COPY terms TO '{(dest / 'terms.parquet').as_posix()}' "
            "(FORMAT PARQUET, COMPRESSION ZSTD)"
        )
    finally:
        con.close()


def _geojson_member(zf: zipfile.ZipFile, zip_name: str) -> str:
    """zip 内の UTF-8 GeoJSON のメンバー名（Shapefile の .dbf は Shift-JIS）。"""
    for info in zf.infolist():
        if not info.is_dir() and _decode_member_name(info.filename).endswith(
            ".geojson"
        ):
            return info.filename
    raise SystemExit(f"no geojson in {zip_name}")


def _convert_prefecture(
    district: SchoolDistrict,
    zip_path: Path,
    tmp_dir: Path,
    parquet_path: Path,
    known_codes: set[str],
    open_codes: set[str],
) -> tuple[int, int]:
    """zip 内のオープンデータ公開の市区町村を 1 つの Parquet にまとめる。

    ジオメトリは WKB (BLOB) で保存し、dbt 側で ST_GeomFromWKB で復元する。
    一時ファイルに書き出してからリネームすることで、中断時に不完全な
    Parquet が変換済みとして残らないようにする。
    戻り値は (取り込んだ市区町村数, 利用条件により除外した市区町村数)。
    """
    geojson_path = tmp_dir / f"{district.key}.geojson"
    with zipfile.ZipFile(zip_path) as zf:
        with zf.open(_geojson_member(zf, zip_path.name)) as src:
            with open(geojson_path, "wb") as dst:
                shutil.copyfileobj(src, dst, 1024 * 1024)

    prefix = district.attribute_prefix
    con = duckdb.connect()
    try:
        con.execute("INSTALL spatial; LOAD spatial;")
        con.execute(
            f"CREATE TEMP TABLE source AS "
            f"SELECT * FROM ST_Read('{geojson_path.as_posix()}')"
        )
        codes = {
            row[0]
            for row in con.execute(
                f"SELECT DISTINCT {prefix}_001 FROM source"
            ).fetchall()
        }
        unknown = sorted(codes - known_codes)
        if unknown:
            raise SystemExit(f"municipalities missing from terms of use: {unknown}")

        included = codes & open_codes
        con.execute(CREATE_TABLE_SQL)
        con.execute(
            f"""
            INSERT INTO school_district
            SELECT
                CAST({prefix}_001 AS VARCHAR),
                CAST({prefix}_002 AS VARCHAR),
                CAST({prefix}_003 AS VARCHAR),
                CAST({prefix}_004 AS VARCHAR),
                CAST({prefix}_005 AS VARCHAR),
                ST_AsWKB(geom)
            FROM source
            WHERE {prefix}_001 IN (SELECT UNNEST(?::VARCHAR[]))
            """,
            [sorted(included)],
        )

        tmp_path = parquet_path.with_suffix(".parquet.tmp")
        con.execute(
            f"COPY school_district TO '{tmp_path.as_posix()}' "
            "(FORMAT PARQUET, COMPRESSION ZSTD)"
        )
        tmp_path.rename(parquet_path)
    finally:
        con.close()
        geojson_path.unlink(missing_ok=True)

    return len(included), len(codes - open_codes)


def download_school_district(dest_dir: str) -> None:
    """通学区域データ（オープンデータ公開の市区町村のみ）を Parquet 化する。

    小学校区・中学校区それぞれについて都道府県ごとに zip をダウンロードし、
    オープンデータ公開の市区町村だけを Parquet にまとめる。zip・GeoJSON は
    変換後すぐ削除する。変換済みの都道府県はスキップするため、途中で中断しても
    再実行で続きから処理できる（冪等）。
    """
    dest = Path(dest_dir)
    tmp_dir = dest / "tmp"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    terms = _load_terms(dest)
    known_codes = {t["admin_code"] for t in terms}
    _write_terms(dest, terms)

    for district in DISTRICTS:
        parquet_dir = dest / district.key
        parquet_dir.mkdir(parents=True, exist_ok=True)
        open_codes = _open_codes(terms, district)

        for pref in _prefectures():
            name = f"{district.identifier}_{pref}"
            parquet_path = parquet_dir / f"{pref}.parquet"
            if parquet_path.exists():
                logger.info(f"  skip {name} (already converted)")
                continue

            zip_path = tmp_dir / f"{name}_GML.zip"
            logger.info(f"  downloading {name}...")
            try:
                _download(district.url_template.format(pref=pref), zip_path)
            except HTTPError as e:
                if e.code == 404:
                    logger.info(f"  skip {name} (not found)")
                    continue
                raise

            try:
                included, excluded = _convert_prefecture(
                    district, zip_path, tmp_dir, parquet_path, known_codes, open_codes
                )
            finally:
                zip_path.unlink(missing_ok=True)

            logger.info(
                f"  {name}: {included} municipalities converted, "
                f"{excluded} excluded by terms of use"
            )

    logger.info(f"  school district data ready in {dest}")
